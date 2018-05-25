"""
    Module for auxiliary functions
    Author: i.zemskov
    Date: 17.05.2018
"""

from openpyxl import load_workbook
import pandas as pd
import datetime as dt


def compare_str_fields(f1, f2):

    """
    Compares values of two string fields in dataset
    :param f1: name of the first field
    :param f2: name of the second field
    :return: comparison result (boolean)
    """

    return str(f1).lower().strip() == str(f2).lower().strip()


def compare_str_series(s1, s2):

    """
    Compares values of two series
    :param s1: 1st series
    :param s2: 2nd series
    :return: result of comparison (boolean generator)
    """

    for f1, f2 in zip(s1, s2):
        yield compare_str_fields(f1, f2)


def to_spreadsheet(df, workbook_path, sheet_name):

    """
    Exports specified data to the excel spreadsheet file on specified sheet
    :param df: data to export
    :param workbook_path: path to the target excel spreadsheet file
    :param sheet_name: name of the target sheet
    """

    book = load_workbook(workbook_path)
    sheet = book[sheet_name]
    index = book.index(sheet)
    book.remove(sheet)
    book.create_sheet(sheet_name, index)
    writer = pd.ExcelWriter(workbook_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # export data to recreated worksheet
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    writer.close()


def from_ts_to_date(ts):

    """
    Converts timestamp object to datetime.date object
    :param ts: timestamp object to convert
    :return: converted datetime.date object
    """

    return ts.to_datetime().date()


def excel_date(date1):

    """
    Converts datetime.datetime to excel date value
    :param date1: datetime.datetime value to convert
    :return: converted value
    """

    temp = dt.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)
